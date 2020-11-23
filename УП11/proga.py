from interface import *
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidgetItem as QTW, QMessageBox
from PyQt5.QtWidgets import QMessageBox as mb
import sqlite3
from tkinter import *
import docx
import openpyxl
import os


class MyWin(QtWidgets.QMainWindow):

    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.pushButton_1.clicked.connect(self.MyFunction)
        self.ui.pushButton_6.clicked.connect(self.MyFunction2)
        self.ui.pushButton_11.clicked.connect(self.MyFunction3)
        self.ui.pushButton_2.clicked.connect(self.udal)
        self.ui.pushButton_7.clicked.connect(self.udal2)
        self.ui.pushButton_12.clicked.connect(self.udal3)
        self.ui.pushButton_3.clicked.connect(self.dobav)
        self.ui.pushButton_8.clicked.connect(self.dobav2)
        self.ui.pushButton_13.clicked.connect(self.dobav3)
        self.ui.pushButton_4.clicked.connect(self.cor)
        self.ui.pushButton_9.clicked.connect(self.cor2)
        self.ui.pushButton_14.clicked.connect(self.cor3)
        self.ui.pushButton_5.clicked.connect(self.popk)
        self.ui.pushButton_10.clicked.connect(self.popk2)
        self.ui.pushButton_15.clicked.connect(self.popk3)
        self.ui.pushButton_leg.clicked.connect(self.legend)
        self.ui.action.triggered.connect(self.oprg)
        self.ui.action_2.triggered.connect(self.oav)
        self.ui.action_3.triggered.connect(self.instr)
        self.ui.action_4.triggered.connect(self.close)
        self.ui.pushButton16.clicked.connect(self.zapr)
        self.ui.action_5.triggered.connect(self.vivvf)
        self.ui.action_6.triggered.connect(self.admin)
        self.ui.action_7.triggered.connect(self.polz)
        self.ui.pushButton_2.setEnabled(False)
        self.ui.pushButton_3.setEnabled(False)
        self.ui.pushButton_4.setEnabled(False)
        self.ui.pushButton_7.setEnabled(False)
        self.ui.pushButton_8.setEnabled(False)
        self.ui.pushButton_9.setEnabled(False)
        self.ui.pushButton_12.setEnabled(False)
        self.ui.pushButton_13.setEnabled(False)
        self.ui.pushButton_14.setEnabled(False)

    def registrat(self, event):
        root.destroy()
        text, ok = QtWidgets.QInputDialog.getText(self, 'Регистрация',
                                                  'Введите новый пароль (без пробелов)')
        if ok:
            clc = text
            clc = str(clc).replace(" ", "")
            print(clc)
            with open('password.txt') as datfile:
                texts = datfile.read()
                texts = texts.split(' ')
            for i in texts:
                if clc == i:
                    mb.about(self, 'Ошибка', 'Такой пароль уже существует')
                    return

            my_filea = open('password.txt', 'a')
            clc = text
            for xxx in clc:
                if xxx == ' ':
                    mb.about(self, 'Ошибка', 'При вводе не допустимы пробелы, все пробелы удалены.')
                else:
                    my_filea.write(xxx)
            my_filea.write(' ')

    def parolll(self, event):
        root.destroy()
        text, ok = QtWidgets.QInputDialog.getText(self, 'Вход',
                                                  'Введите пароль')
        if ok:
            clc = text
            with open('password.txt') as datfile:
                text = datfile.read()
                text = text.split(' ')
                for i in text:
                    if clc == i:
                        self.ui.pushButton_2.setEnabled(True)
                        self.ui.pushButton_3.setEnabled(True)
                        self.ui.pushButton_4.setEnabled(True)
                        self.ui.pushButton_7.setEnabled(True)
                        self.ui.pushButton_8.setEnabled(True)
                        self.ui.pushButton_9.setEnabled(True)
                        self.ui.pushButton_12.setEnabled(True)
                        self.ui.pushButton_13.setEnabled(True)
                        self.ui.pushButton_14.setEnabled(True)
                        mb.about(self, 'Успешно', 'Вы вошли в систему как администратор.')

    def admin(self):
        print(1)
        global root
        root = Tk()
        root.geometry('400x100+800+300')
        root.title('Пароль')
        print(1)
        fra = Frame(root, width=400, height=100)
        fra.grid(row=0, column=0)
        print(1)
        but = Button(fra, text='Зарегистрироваться', height=2, font='Arial 10')
        but.bind('<Button-1>', self.registrat)
        but.place(x=10, y=20)
        but2 = Button(fra, text='Войти', height=2, font='Arial 12')
        but2.bind('<Button-1>', self.parolll)
        but2.place(x=250, y=10)
        print(1)
        root.mainloop()

    def polz(self):
        self.ui.pushButton_2.setEnabled(False)
        self.ui.pushButton_3.setEnabled(False)
        self.ui.pushButton_4.setEnabled(False)
        self.ui.pushButton_7.setEnabled(False)
        self.ui.pushButton_8.setEnabled(False)
        self.ui.pushButton_9.setEnabled(False)
        self.ui.pushButton_12.setEnabled(False)
        self.ui.pushButton_13.setEnabled(False)
        self.ui.pushButton_14.setEnabled(False)

    def legend(self):
        mb.about(self, 'Список запросов',
                 '1.В каком кабинете находится данное оборудование?\n2.Сколько оборудования находится в данном кабинете?\n3.На каком этаже находится этот кабинет?\n4.Контактный телефон ответственного за данный кабинет?\n5.Вывести ФИО сотрудников и кабинеты за которые они несут ответственность.\n6.Сколько кабинетов на данном этаже?\n7.Название кабинета в котором находится данное оборудование?\n8.Сколько сотрудников ответственны за данный кабинет?\n9.Сколько оборудования находится на данном этаже?\n10.ФИО ответственного за этот кабинет?')

    def oprg(self):
        mb.about(self, 'О программе',
                 'Программа предназначена для взаимодействия с БД АиРЭО. Для показа, ввода,удаления и т.д данных.')

    def oav(self):
        mb.about(self, 'Об Авторе', 'Миронов Даниил Сергеевич - курсант 431 группы Троицкого АТК - филиала МГТУ ГА')

    def instr(self):
        mb.about(self, 'Инструкция к применению', 'Когда программа запускается, нас встречает интерфейс'
                                                  ', на котором находятся кнопки , выполняющие соответствующие им функции. '
                                                  'Над полем таблицы присутствует выбор активной на данной момент таблицы. '
                                                  'Над программой есть меню, в котором можно узнать краткую информацию обо всем интересующем.')

    def vivvf(self):

        numTab = self.ui.tabWidget.currentIndex()
        wb = openpyxl.Workbook()
        wb.create_sheet(title = 'Первый лист',index = 0)
        sheet = wb ['Первый лист']
        if numTab == 0:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            cur.execute('SELECT * FROM Сотрудник')
            res = cur.fetchall()
            tble = []
            my_file = open('Сотрудники.txt', 'w')
            for j in range(len(res)):
                spk = 4
                for i in range(len(res[0])):
                    tb = self.ui.tableWidget.item(j, i).text()
                    my_file.write(tb + ' ')
                    spk = spk - 1
                    if spk == 0:
                        my_file.write('\n')
            for j in range(1,len(res)+1):
                for i in range (1,len(res[0])+1):
                    tb = self.ui.tableWidget.item(j-1,i-1).text()
                    cell = sheet.cell(row = j, column = i)
                    cell.value =tb
            wb.save('Кабинет.xlsx')
            mb.about(self,"Готово","Данные таблицы Кабинет занесены в excel и txt файлы")
            conn.commit()
            conn.close()
            my_file.close()

        if numTab == 1:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            cur.execute('SELECT * FROM Кабинет')
            res = cur.fetchall()
            tble = []
            my_file = open('Кабинет.txt', 'w')
            for j in range(len(res)):
                spk = 4
                for i in range(len(res[0])):
                    tb = self.ui.tableWidget_2.item(j, i).text()
                    my_file.write(tb + ' ')
                    spk = spk - 1
                    if spk == 0:
                        my_file.write('\n')
            conn.commit()
            conn.close()
            my_file.close()
        if numTab == 2:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            cur.execute('SELECT * FROM Оборудование')
            res = cur.fetchall()
            tble = []
            my_file = open('Оборудование.txt', 'w')
            for j in range(len(res)):
                spk = 2
                for i in range(len(res[0])):
                    tb = self.ui.tableWidget_3.item(j, i).text()
                    my_file.write(tb + ' ')
                    spk = spk - 1
                    if spk == 0:
                        my_file.write('\n')
            conn.commit()
            conn.close()
            my_file.close()

    def zapr(self):

        scitv = int(self.ui.spinBox.text())
        if scitv == 1:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 1',
                                                      'Введите название оборудования.')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute('SELECT * FROM Оборудование WHERE Название = ' + '"' + (clc) + '"')
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()
                self.ui.tableWidget_4.setHorizontalHeaderLabels(['НомерКабинета', 'НазваниеОборудования'])



        elif scitv == 2:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 2', 'Введите номер кабинета')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute('SELECT * FROM Кабинет WHERE НомерКабинета = ' + str(clc))
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()
                self.ui.tableWidget_4.setHorizontalHeaderLabels(
                    ['НомерКабинета', 'Название', 'Этаж', 'НомерОтветственногоЗаКабинет'])

        elif scitv == 3:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 3', 'Введите номер кабинета ')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute(
                    'SELECT Сотрудник.КонтактныйНомер , Сотрудник.ФИО From Сотрудник, Кабинет Where Кабинет.НомерОтветственногоЗаКабинет = Сотрудник.id_сотрудника AND 	Кабинет.НомерКабинета = ' + str(
                        clc))
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()
                self.ui.tableWidget_4.setHorizontalHeaderLabels(['НомерКабинета', 'НазваниеОборудования'])

        elif scitv == 4:

            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            print(1)
            cur.execute(
                'Select Сотрудник.ФИО, Кабинет.НомерКабинета From Сотрудник, Кабинет Where Кабинет.НомерОтветственногоЗаКабинет = Сотрудник.id_сотрудника')
            print(1)
            result = cur.fetchall()
            res = result
            self.prosmsotrzapr1(res, self.ui.tableWidget_4)
            conn.commit()
            conn.close()
        elif scitv == 5:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 5', 'Введите номер кабинета')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute(
                    'Select Сотрудник.ФИО From Сотрудник,Кабинет Where Кабинет.НомерОтветственногоЗаКабинет = Сотрудник.id_сотрудника and Кабинет.НомерКабинета = ' + str(
                        clc))
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()

        elif scitv == 6:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 6', 'Введите номер этажа: ')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute(
                    'Select Count(НомерКабинета) From Кабинет Where Этаж = ' + str(
                        clc))
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()
        elif scitv == 7:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 7', 'Введите название оборудования: ')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute(
                    'Select Кабинет.Название From Кабинет, Оборудование Where Кабинет.НомерКабинета = Оборудование.НомерКабинета and Оборудование.Название = ' + '"' + (
                        clc) + '"')
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()
        elif scitv == 8:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 8', 'Введите номер кабинета: ')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute(
                    'Select Count(НомерОтветственногоЗаКабинет) From Кабинет Where НомерКабинета = ' + str(
                        clc))
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()
        elif scitv == 9:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 9', 'Введите номер этажа: ')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute(
                    'Select Count(Оборудование.Название) From Оборудование, Кабинет Where Кабинет.НомерКабинета = Оборудование.НомерКабинета and Кабинет.Этаж = ' + str(
                        clc))
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()
        elif scitv == 10:
            text, ok = QtWidgets.QInputDialog.getText(self, 'Запрос 10', 'Введите номер кабинета: ')
            if ok:
                clc = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                print(1)
                cur.execute(
                    'Select Сотрудник.ФИО From Сотрудник, Кабинет Where Кабинет.НомерОтветственногоЗаКабинет = Сотрудник.id_сотрудника and Кабинет.НомерКабинета = ' + str(
                        clc))
                print(1)
                result = cur.fetchall()
                res = result
                self.prosmsotrzapr1(res, self.ui.tableWidget_4)
                conn.commit()
                conn.close()

    def prosmsotrzapr1(self, res, tbl):
        k = 0
        tbl.setRowCount(len(res))
        tbl.setColumnCount(len(res[0]))
        for j in range(len(res)):
            k += 1
            for i in range(len(res[0])):
                tbl.setItem(j, i, QTW(str(res[j][i])))
                self.ui.label_2.setText('Количество записей:' + str(k))

    def MyFunction(self):
        conn = sqlite3.connect('aireo.db')
        cur = conn.cursor()
        cur.execute('SELECT * FROM Сотрудник')
        result = cur.fetchall()
        res = result
        self.prosmsotr(res, self.ui.tableWidget)
        conn.commit()
        conn.close()

    def prosmsotr(self, res, tbl):
        k = 0
        tbl.setRowCount(len(res))
        tbl.setColumnCount(len(res[0]))
        for j in range(len(res)):
            k += 1
            for i in range(len(res[0])):
                tbl.setItem(j, i, QTW(str(res[j][i])))
                self.ui.label_2.setText('Количество записей:' + str(k))

    def MyFunction2(self):
        conn = sqlite3.connect('aireo.db')
        cur = conn.cursor()
        cur.execute('SELECT * FROM Кабинет')
        result = cur.fetchall()
        res = result
        self.prosmsotr(res, self.ui.tableWidget_2)
        conn.commit()
        conn.close()

    def prosmsotr2(self, res, tbl):
        sqfsq = 0
        tbl.setRowCount(len(res))
        tbl.setColumnCount(len(res[0]))
        for j in range(len(res)):
            sqfsq += 1
            for i in range(len(res[0])):
                tbl.setItem(j, i, QTW(str(res[j][i])))
                self.ui.label_2.setText('Количество записей:' + str(sqfsq))

    def MyFunction3(self):
        conn = sqlite3.connect('aireo.db')
        cur = conn.cursor()
        cur.execute('SELECT * FROM Оборудование')
        result = cur.fetchall()
        res = result
        self.prosmsotr(res, self.ui.tableWidget_3)
        conn.commit()
        conn.close()

    def prosmsotr3(self, res, tbl):
        k = 0
        tbl.setRowCount(len(res))
        tbl.setColumnCount(len(res[0]))
        for j in range(len(res)):
            k += 1
            for i in range(len(res[0])):
                tbl.setItem(j, i, QTW(str(res[j][i])))
                self.ui.label_2.setText('Количество записей:' + str(k))

    def udal(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Удаление', 'Введите id сотрудника для удаления данных о нем ')
        if ok:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            rest = text
            cur.execute('DELETE FROM Сотрудник WHERE Id_сотрудника = ?', (rest,))

            conn.commit()
            conn.close()

    def udal2(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Удаление', 'Введите номер кабинета для удаления ')
        if ok:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            rest = text
            print(1)
            cur.execute("DELETE FROM Кабинет WHERE НомерКабинета = ?", (rest,))
            print(1)
            conn.commit()
            conn.close()

    def udal3(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Удаление', 'Введите название оборудования для удаления ')
        if ok:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            rest = text
            cur.execute('DELETE FROM Оборудование WHERE Название = ?', (rest,))

            conn.commit()
            conn.close()

    def dobav(self):

        text, ok = QtWidgets.QInputDialog.getText(self, 'Добавление', 'Введите новые данные через запятую.')
        if ok:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            rest = text
            pp = []
            rest = rest.split(',')
            for i in rest:
                pp.append(i)
            pp = tuple(pp)

            cur.execute('INSERT INTO Сотрудник (id_сотрудника, ФИО, КонтактныйНомер, ДатаРождения) VALUES (?, ?, ?, ?)',
                        pp)

            conn.commit()
            conn.close()

    def dobav2(self):

        text, ok = QtWidgets.QInputDialog.getText(self, 'Добавление', 'Введите новые данные через запятую.')
        if ok:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            rest = text
            pp = []
            rest = rest.split(',')
            for i in rest:
                pp.append(i)
            pp = tuple(pp)

            cur.execute(
                'INSERT INTO Кабинет (НомерКабинета, Название, Этаж, НомерОтветственногоЗаКабинет) VALUES (?, ?, ?, ?)',
                pp)

            conn.commit()
            conn.close()

    def dobav3(self):

        text, ok = QtWidgets.QInputDialog.getText(self, 'Добавление', 'Введите новые данные через запятую.')
        if ok:
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            rest = text
            pp = []
            rest = rest.split(',')
            for i in rest:
                pp.append(i)
            pp = tuple(pp)

            cur.execute('INSERT INTO Оборудование (НомерКабинета, Название) VALUES (?, ?)', pp)

            conn.commit()
            conn.close()

    def cor(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Изменение', 'Введите номер строки ')
        if ok:
            nst = text
            text, ok = QtWidgets.QInputDialog.getText(self, 'Изменение', 'Введите новую строчку через запятую')
            if ok:
                vms = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                pp = []
                vms = vms.split(',')
                for i in vms:
                    pp.append(i)
                pp = tuple(pp)

                ids = self.ui.tableWidget.item(int(nst) - 1, 0).text()

                cur.execute(
                    'UPDATE Сотрудник SET id_сотрудника = ?, ФИО = ?, КонтактныйНомер = ?, ДатаРождения = ? WHERE id_сотрудника = ' + str(
                        ids), pp)
                conn.commit()
                conn.close()

    def cor2(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Изменение', 'Введите номер строки ')
        if ok:
            nst = text
            text, ok = QtWidgets.QInputDialog.getText(self, 'Изменение', 'Введите новую строчку через запятую')
            if ok:
                vms = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                pp = []
                vms = vms.split(',')
                for i in vms:
                    pp.append(i)
                pp = tuple(pp)

                ids = self.ui.tableWidget_2.item(int(nst) - 1, 0).text()

                cur.execute(
                    'UPDATE Кабинет SET НомерКабинета = ?, Название = ?, Этаж = ?, НомерОтветственногоЗаКабинет = ? WHERE НомерКабинета = ' + str(
                        ids), pp)
                conn.commit()
                conn.close()

    def cor3(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Изменение', 'Введите номер строки ')
        if ok:
            nst = text
            text, ok = QtWidgets.QInputDialog.getText(self, 'Изменение', 'Введите новую строчку через запятую')
            if ok:
                vms = text
                conn = sqlite3.connect('aireo.db')
                cur = conn.cursor()
                pp = []
                vms = vms.split(',')
                for i in vms:
                    pp.append(i)
                pp = tuple(pp)

                ids = self.ui.tableWidget_3.item(int(nst) - 1, 0).text()
                ids2 = self.ui.tableWidget_3.item(int(nst) - 1, 1).text()
                print(ids2)
                print(1)
                # cur.execute('UPDATE Оборудование SET НомерКабинета = ?, Название = ? WHERE НомерКабинета = '+ str(ids) +'AND Название = '+str(ids2), pp)
                cur.execute('UPDATE Оборудование SET НомерКабинета = ?, Название = ? WHERE НомерКабинета = ' + str(
                    ids) + ' AND Название = ' + '"' + ids2 + '"', pp)
                print(1)
                conn.commit()
                conn.close()

    def popk(self, tbl):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Поиск по ключу', 'Введите значение ключевого поля')
        if ok:
            clc = text
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            print(1)
            cur.execute('SELECT * FROM Сотрудник WHERE id_сотрудника = ' + str(clc))
            print(1)
            result = cur.fetchall()
            res = result
            self.prosmsotr(res, self.ui.tableWidget)

            conn.commit()
            conn.close()

    def popk2(self, tbl):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Поиск по ключу', 'Введите значение ключевого поля')
        if ok:
            clc = text
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            print(1)
            cur.execute('SELECT * FROM Кабинет WHERE НомерКабинета = ' + str(clc))
            print(1)
            result = cur.fetchall()
            res = result
            self.prosmsotr2(res, self.ui.tableWidget_2)

            conn.commit()
            conn.close()

    def popk3(self, tbl):
        text, ok = QtWidgets.QInputDialog.getText(self, 'Поиск по ключу', 'Введите значение ключевого поля')
        if ok:
            clc = text
            conn = sqlite3.connect('aireo.db')
            cur = conn.cursor()
            print(1)
            cur.execute('SELECT * FROM Оборудование WHERE НомерКабинета = ' + str(clc))
            print(1)
            result = cur.fetchall()
            res = result
            self.prosmsotr3(res, self.ui.tableWidget_3)

            conn.commit()
            conn.close()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin()
    myapp.show()
    sys.exit(app.exec_())
